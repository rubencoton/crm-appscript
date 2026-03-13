#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import collections
import datetime as dt
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.parse
import urllib.request
from typing import Any

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJ = os.path.join(ROOT, "crm-ayudas-subvenciones")
REPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
CLASPRC = os.path.join(os.environ.get("USERPROFILE", ""), ".clasprc.json")


def norm(s: str) -> str:
    s = (s or "").strip().upper()
    for a, b in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ü", "U"), ("Ñ", "N")]:
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s)


def a1q(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def n2c(n: int) -> str:
    if n <= 0:
        return "A"
    out = []
    while n:
        n, r = divmod(n - 1, 26)
        out.append(chr(65 + r))
    return "".join(reversed(out))


def c2n(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def http_json(method: str, url: str, token: str | None = None, data: dict | None = None) -> dict:
    h = {"Accept": "application/json"}
    b = None
    if token:
        h["Authorization"] = "Bearer " + token
    if data is not None:
        b = json.dumps(data).encode("utf-8")
        h["Content-Type"] = "application/json"
    req = urllib.request.Request(url=url, method=method, headers=h, data=b)
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        txt = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url} :: {txt}") from e


def http_bytes(method: str, url: str, token: str | None = None) -> bytes:
    h = {"Accept": "*/*"}
    if token:
        h["Authorization"] = "Bearer " + token
    req = urllib.request.Request(url=url, method=method, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            return r.read()
    except urllib.error.HTTPError as e:
        txt = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url} :: {txt}") from e


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def refresh_token(clasprc: str) -> str:
    cfg = load_json(clasprc)
    cid = csec = rt = ""
    if cfg.get("token") and cfg.get("oauth2ClientSettings"):
        cid = str(cfg["oauth2ClientSettings"].get("clientId", ""))
        csec = str(cfg["oauth2ClientSettings"].get("clientSecret", ""))
        rt = str(cfg["token"].get("refresh_token", ""))
    elif cfg.get("tokens", {}).get("default"):
        t = cfg["tokens"]["default"]
        cid = str(t.get("client_id", ""))
        csec = str(t.get("client_secret", ""))
        rt = str(t.get("refresh_token", ""))
    if not (cid and csec and rt):
        raise RuntimeError("No OAuth en .clasprc.json")
    tok = http_json("POST", "https://oauth2.googleapis.com/token", data={
        "client_id": cid, "client_secret": csec, "refresh_token": rt, "grant_type": "refresh_token"
    })
    if not tok.get("access_token"):
        raise RuntimeError("Refresh OAuth sin access_token")
    return tok["access_token"]


def parse_used_range(raw: str, values: list) -> tuple[int, int, str]:
    if values:
        rr = len(values)
        cc = max((len(r) for r in values), default=0)
        return rr, cc, f"A1:{n2c(max(cc,1))}{max(rr,1)}"
    if "!" not in raw:
        return 0, 0, "A1"
    a1 = raw.split("!", 1)[1]
    m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", a1)
    if not m:
        return 0, 0, a1
    if m.group(3):
        return int(m.group(4)), c2n(m.group(3)), a1
    return int(m.group(2)), c2n(m.group(1)), a1


def sheet_meta(token: str, sid: str) -> dict:
    fields = ("spreadsheetId,properties(title,locale,timeZone),"
              "sheets(properties(sheetId,title,index,hidden,gridProperties(rowCount,columnCount,frozenRowCount,frozenColumnCount)),"
              "merges,basicFilter,filterViews,conditionalFormats,protectedRanges)")
    q = urllib.parse.quote(fields, safe=",()")
    return http_json("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{sid}?includeGridData=false&fields={q}", token)


def sheet_values(token: str, sid: str, names: list[str]) -> dict:
    params = [("valueRenderOption", "FORMULA"), ("dateTimeRenderOption", "SERIAL_NUMBER")]
    for n in names:
        params.append(("ranges", a1q(n)))
    qs = urllib.parse.urlencode(params, doseq=True, safe="'!")
    return http_json("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{sid}/values:batchGet?{qs}", token)


def sheet_grid(token: str, sid: str, a1: str) -> dict:
    fields = ("sheets(properties(title),data(rowData(values("
              "formattedValue,userEnteredValue,effectiveValue,userEnteredFormat(numberFormat,horizontalAlignment,verticalAlignment),"
              "dataValidation))))")
    qs = urllib.parse.urlencode({"includeGridData": "true", "ranges": a1, "fields": fields}, safe="'!,()")
    return http_json("GET", f"https://sheets.googleapis.com/v4/spreadsheets/{sid}?{qs}", token)


def script_project(token: str, script_id: str) -> dict:
    return http_json("GET", f"https://script.googleapis.com/v1/projects/{script_id}", token)


def script_content(token: str, script_id: str) -> dict:
    return http_json("GET", f"https://script.googleapis.com/v1/projects/{script_id}/content", token)


def export_sheet_xlsx(token: str, spreadsheet_id: str, out_path: str) -> str:
    url = (
        f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}/export"
        "?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    data = http_bytes("GET", url, token)
    with open(out_path, "wb") as f:
        f.write(data)
    return out_path


def sha(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8", errors="ignore")).hexdigest()


def audit_code(project_dir: str, token: str, script_id: str) -> tuple[dict, list[dict]]:
    findings = []
    files = [os.path.join(project_dir, "Code.js"), os.path.join(project_dir, "DecisionInstantanea.js")]
    stats, syntax = {}, {}
    fn_glob = collections.Counter()
    mojis = {}
    for fp in files:
        name = os.path.basename(fp)
        if not os.path.exists(fp):
            findings.append({"severity": "CRITICAL", "area": "CODE_LOCAL", "title": "Archivo ausente",
                             "detail": name, "evidence": {"file": fp}, "recommendation": "Restaurar archivo."})
            continue
        txt = open(fp, "r", encoding="utf-8", errors="replace").read()
        p = subprocess.run(["node", "--check", fp], capture_output=True, text=True, shell=False)
        syntax[name] = {"ok": p.returncode == 0, "stderr": p.stderr.strip(), "stdout": p.stdout.strip()}
        if p.returncode != 0:
            findings.append({"severity": "CRITICAL", "area": "CODE_SYNTAX", "title": "Error de sintaxis",
                             "detail": name, "evidence": syntax[name], "recommendation": "Corregir JS."})
        funcs = re.findall(r"\bfunction\s+([A-Za-z0-9_]+)\s*\(", txt)
        for f in funcs:
            fn_glob[f] += 1
        moj = len(re.findall(r"(ðŸ|âš|âœ|Ã|Â)", txt))
        mojis[name] = moj
        stats[name] = {"lines": txt.count("\n") + 1, "bytes": len(txt.encode("utf-8")), "sha256": sha(txt), "functions": len(funcs), "mojibakeTokens": moj}
    dups = [k for k, v in fn_glob.items() if v > 1]
    if dups:
        findings.append({"severity": "MEDIUM", "area": "CODE_QUALITY", "title": "Funciones duplicadas",
                         "detail": f"{len(dups)} nombres repetidos", "evidence": {"duplicates": dups[:40]}, "recommendation": "Revisar colisiones."})
    if sum(mojis.values()) > 0:
        findings.append({"severity": "HIGH", "area": "CODE_UX", "title": "Texto corrupto (mojibake)",
                         "detail": "Se ven tokens rotos de codificación", "evidence": {"byFile": mojis}, "recommendation": "Normalizar UTF-8."})
    secret_hits = []
    pats = [r"AIza[0-9A-Za-z\-_]{20,}", r"sk-[A-Za-z0-9]{20,}", r"(?i)api[_-]?key\s*[:=]\s*['\"][^'\"]{8,}['\"]"]
    for fp in files:
        if not os.path.exists(fp):
            continue
        txt = open(fp, "r", encoding="utf-8", errors="replace").read()
        for pat in pats:
            for m in re.finditer(pat, txt):
                secret_hits.append({"file": os.path.basename(fp), "pattern": pat, "excerpt": m.group(0)[:120]})
    if secret_hits:
        findings.append({"severity": "CRITICAL", "area": "CODE_SECURITY", "title": "Posibles secretos en código",
                         "detail": f"{len(secret_hits)} coincidencias", "evidence": {"hits": secret_hits[:20]}, "recommendation": "Mover secretos a propiedades."})
    remote = script_content(token, script_id)
    rf = {(f.get("name", "") + "." + {"SERVER_JS": "js", "JSON": "json", "HTML": "html"}.get(f.get("type", ""), "")).strip(".")
          for f in remote.get("files", [])}
    miss = [x for x in ["appsscript.json", "Code.js", "DecisionInstantanea.js"] if x not in rf]
    if miss:
        findings.append({"severity": "HIGH", "area": "CODE_SYNC", "title": "Remoto incompleto",
                         "detail": "Faltan archivos en Apps Script remoto", "evidence": {"missing": miss}, "recommendation": "Hacer clasp push -f."})
    return {"stats": stats, "syntax": syntax, "remoteFiles": sorted(list(rf))}, findings


def audit_sheet(token: str, spreadsheet_id: str) -> tuple[dict, list[dict]]:
    findings = []
    md = sheet_meta(token, spreadsheet_id)
    sheets = md.get("sheets", [])
    names = [s.get("properties", {}).get("title", "") for s in sheets]
    vb = sheet_values(token, spreadsheet_id, names)
    vmap = {}
    for vr in vb.get("valueRanges", []):
        n = vr.get("range", "").split("!")[0].strip("'")
        vals = vr.get("values", []) or []
        rr, cc, a1 = parse_used_range(vr.get("range", ""), vals)
        vmap[n] = {"values": vals, "usedRows": rr, "usedCols": cc, "usedRange": a1}

    structure, feats, fmt = [], {}, {}
    for s in sheets:
        p = s.get("properties", {})
        n = p.get("title", "")
        gp = p.get("gridProperties", {})
        vv = vmap.get(n, {"usedRows": 0, "usedCols": 0, "usedRange": "A1", "values": []})
        structure.append({
            "title": n, "sheetId": p.get("sheetId"), "index": p.get("index"), "hidden": p.get("hidden", False),
            "rowsConfigured": gp.get("rowCount"), "colsConfigured": gp.get("columnCount"),
            "frozenRows": gp.get("frozenRowCount", 0), "frozenCols": gp.get("frozenColumnCount", 0),
            "usedRows": vv["usedRows"], "usedCols": vv["usedCols"], "usedRange": vv["usedRange"],
        })
        feats[n] = {
            "merges": len(s.get("merges", []) or []),
            "basicFilter": bool(s.get("basicFilter")),
            "filterViews": len(s.get("filterViews", []) or []),
            "conditionalRules": len(s.get("conditionalFormats", []) or []),
            "protectedRanges": len(s.get("protectedRanges", []) or []),
        }
        if vv["usedRows"] <= 0 or vv["usedCols"] <= 0:
            fmt[n] = {"cellsScanned": 0, "validations": 0, "errorsInCells": 0, "validationsByCol": {}}
            continue
        a1 = f"{a1q(n)}!A1:{n2c(vv['usedCols'])}{vv['usedRows']}"
        gd = sheet_grid(token, spreadsheet_id, a1)
        row_data = (((gd.get("sheets", [{}])[0]).get("data", [{}])[0]).get("rowData", [])) or []
        validations = 0
        v_by_col = collections.Counter()
        errs = 0
        scanned = 0
        for row in row_data:
            for ci, c in enumerate(row.get("values", []) or [], start=1):
                scanned += 1
                if c.get("dataValidation"):
                    validations += 1
                    v_by_col[ci] += 1
                if "errorValue" in (c.get("effectiveValue", {}) or {}):
                    errs += 1
        fmt[n] = {"cellsScanned": scanned, "validations": validations, "errorsInCells": errs,
                  "validationsByCol": {str(k): v for k, v in v_by_col.items()}}

    cv = vmap.get("CONCURSOS", {}).get("values", [])
    if not cv:
        findings.append({"severity": "CRITICAL", "area": "SHEET_DATA", "title": "CONCURSOS vacío/no legible",
                         "detail": "No hay matriz de datos disponible", "evidence": {}, "recommendation": "Revisar pestaña/permisos."})
    else:
        h = [norm(str(x)) for x in cv[0]]
        rows = cv[1:]
        hm = {h[i]: i for i in range(len(h))}
        exp = ["NOMBRE CONCURSO", "ESTADO", "INSCRIPCION", "FECHA LIMITE INSCRIPCION", "EMAIL"]
        miss = [x for x in exp if x not in hm]
        if miss:
            findings.append({"severity": "HIGH", "area": "SHEET_STRUCTURE", "title": "Cabeceras ausentes",
                             "detail": "Faltan columnas clave", "evidence": {"missing": miss}, "recommendation": "Alinear cabeceras."})
        i_nom = hm.get("NOMBRE CONCURSO", 0); i_est = hm.get("ESTADO", 1); i_ins = hm.get("INSCRIPCION", 2)
        i_fli = hm.get("FECHA LIMITE INSCRIPCION", 3); i_ema = hm.get("EMAIL", 14)
        i_l1 = hm.get("LINK BASES/FORMULARIO", 11); i_l2 = hm.get("WEB / BASES / FORMULARIO ANO ANTERIOR", 12); i_l3 = hm.get("WEB / BASES / FORMULARIO 2 ANOS ANTERIOR", 13)
        ok_est = {"REVISAR", "REVISADO IA", "REVISADO HUMANO", "NUEVO DESCUBRIMIENTO"}
        ok_ins = {"ABIERTA", "CERRADA", "SIN PUBLICAR"}
        bad_est = []; bad_ins = []; bad_mail = []; bad_link = []; bad_name = []; bad_date = []
        dups = collections.defaultdict(list)
        for rn, row in enumerate(rows, start=2):
            v = lambda i: str(row[i]).strip() if i < len(row) else ""
            nom = v(i_nom); est = norm(v(i_est)); ins = norm(v(i_ins)); fl = v(i_fli); em = v(i_ema).lower()
            if not nom and any(str(x).strip() for x in row):
                bad_name.append(rn)
            if nom:
                dups[norm(nom)].append(rn)
            if est and est not in ok_est:
                bad_est.append({"row": rn, "value": est})
            if ins and ins not in ok_ins:
                bad_ins.append({"row": rn, "value": ins})
            if em and "no publicado" not in em and "no localizado" not in em:
                if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", em):
                    bad_mail.append({"row": rn, "value": em})
            for lk in [v(i_l1), v(i_l2), v(i_l3)]:
                if lk and "NO PUBLICADO" not in norm(lk) and "NO LOCALIZADO" not in norm(lk):
                    if not re.match(r"^https?://", lk, re.I):
                        bad_link.append({"row": rn, "value": lk})
            if fl and "ESTIMADO" not in norm(fl):
                if not (re.match(r"^\d+(\.\d+)?$", fl) or re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", fl) or re.search(r"\d{4}-\d{1,2}-\d{1,2}", fl)):
                    bad_date.append({"row": rn, "value": fl})
        dup_groups = [{"nameNorm": k, "rows": v} for k, v in dups.items() if k and len(v) > 1]
        if bad_est:
            findings.append({"severity": "HIGH", "area": "SHEET_DATA", "title": "ESTADO inválido",
                             "detail": f"{len(bad_est)} filas", "evidence": {"examples": bad_est[:15]}, "recommendation": "Corregir catálogo ESTADO."})
        if bad_ins:
            findings.append({"severity": "HIGH", "area": "SHEET_DATA", "title": "INSCRIPCIÓN inválida",
                             "detail": f"{len(bad_ins)} filas", "evidence": {"examples": bad_ins[:15]}, "recommendation": "Corregir catálogo INSCRIPCIÓN."})
        if bad_mail:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Email sospechoso",
                             "detail": f"{len(bad_mail)} filas", "evidence": {"examples": bad_mail[:15]}, "recommendation": "Revisar formato email."})
        if bad_link:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Links inválidos",
                             "detail": f"{len(bad_link)} filas", "evidence": {"examples": bad_link[:15]}, "recommendation": "Añadir https://."})
        if bad_name:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Filas sin nombre",
                             "detail": f"{len(bad_name)} filas", "evidence": {"rows": bad_name[:30]}, "recommendation": "Completar o limpiar filas."})
        if dup_groups:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Posibles duplicados",
                             "detail": f"{len(dup_groups)} grupos", "evidence": {"examples": dup_groups[:15]}, "recommendation": "Depurar duplicados."})
        if bad_date:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Fecha límite no parseable",
                             "detail": f"{len(bad_date)} filas", "evidence": {"examples": bad_date[:15]}, "recommendation": "Estandarizar fecha."})
        dat_rows = max(vmap.get("CONCURSOS", {}).get("usedRows", 0) - 1, 0)
        vbc = {int(k): int(v) for k, v in fmt.get("CONCURSOS", {}).get("validationsByCol", {}).items()}
        weak = []
        for cname, idx in [("ESTADO", i_est + 1), ("INSCRIPCION", i_ins + 1)]:
            if dat_rows > 0 and vbc.get(idx, 0) < int(dat_rows * 0.8):
                weak.append({"column": cname, "index": idx, "validatedCells": vbc.get(idx, 0), "dataRows": dat_rows})
        if weak:
            findings.append({"severity": "HIGH", "area": "SHEET_VALIDATIONS", "title": "Validaciones incompletas",
                             "detail": "Columnas clave sin validación consistente", "evidence": {"columns": weak}, "recommendation": "Aplicar dropdown a toda la matriz."})

    err_sheets = [{"sheet": n, "errorsInCells": x["errorsInCells"]} for n, x in fmt.items() if int(x.get("errorsInCells", 0)) > 0]
    if err_sheets:
        findings.append({"severity": "HIGH", "area": "SHEET_FORMULAS", "title": "Errores de fórmula/valor",
                         "detail": "Hay celdas con error efectivo", "evidence": {"sheets": err_sheets}, "recommendation": "Corregir celdas con #ERROR/#REF."})
    total_prot = sum(feats[n]["protectedRanges"] for n in feats)
    if total_prot == 0:
        findings.append({"severity": "LOW", "area": "SHEET_PROTECTION", "title": "Sin protecciones",
                         "detail": "No hay rangos protegidos", "evidence": {"totalProtectedRanges": 0}, "recommendation": "Proteger celdas críticas."})

    overview = {
        "spreadsheet": {"id": spreadsheet_id, "title": md.get("properties", {}).get("title", ""), "locale": md.get("properties", {}).get("locale", ""), "timeZone": md.get("properties", {}).get("timeZone", ""), "sheetCount": len(sheets)},
        "structure": structure,
        "sheetFeatures": feats,
        "formatsSummary": fmt,
        "valuesPreview": {n: {"usedRange": vmap.get(n, {}).get("usedRange", "A1"), "usedRows": vmap.get(n, {}).get("usedRows", 0), "usedCols": vmap.get(n, {}).get("usedCols", 0), "header": (vmap.get(n, {}).get("values", [[]]) or [[]])[0][:25] if vmap.get(n, {}).get("values") else [], "sampleRows": (vmap.get(n, {}).get("values", [])[1:6])} for n in names},
    }
    return overview, findings


def audit_sheet_xlsx_fallback(token: str, spreadsheet_id: str) -> tuple[dict, list[dict]]:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    findings: list[dict[str, Any]] = []
    tmp = os.path.join(tempfile.gettempdir(), f"crm_ayudas_audit_{spreadsheet_id}.xlsx")
    export_sheet_xlsx(token, spreadsheet_id, tmp)
    wb = load_workbook(tmp, data_only=False)

    try:
        dfile = http_json(
            "GET",
            f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}?fields=id,name,mimeType,modifiedTime",
            token,
        )
        title = dfile.get("name", "")
    except Exception:
        title = ""

    structure, feats, fmt, vprev = [], {}, {}, {}
    concursos_rows: list[list[Any]] = []

    for ws in wb.worksheets:
        dim = ws.calculate_dimension()
        used_rows = int(ws.max_row or 0)
        used_cols = int(ws.max_column or 0)
        structure.append(
            {
                "title": ws.title,
                "sheetId": None,
                "index": wb.sheetnames.index(ws.title),
                "hidden": ws.sheet_state != "visible",
                "rowsConfigured": used_rows,
                "colsConfigured": used_cols,
                "frozenRows": ws.freeze_panes.row - 1 if ws.freeze_panes else 0,
                "frozenCols": ws.freeze_panes.column - 1 if ws.freeze_panes else 0,
                "usedRows": used_rows,
                "usedCols": used_cols,
                "usedRange": dim,
            }
        )

        merges = len(list(ws.merged_cells.ranges))
        cond_rules = len(ws.conditional_formatting)
        has_filter = bool(ws.auto_filter and ws.auto_filter.ref)
        prot = 1 if ws.protection and ws.protection.sheet else 0

        feats[ws.title] = {
            "merges": merges,
            "basicFilter": has_filter,
            "filterViews": 0,
            "conditionalRules": cond_rules,
            "protectedRanges": prot,
        }

        val_count = 0
        val_by_col = collections.Counter()
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                for r in str(dv.sqref).split():
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(r)
                    except Exception:
                        continue
                    cells = max(1, (max_col - min_col + 1) * (max_row - min_row + 1))
                    val_count += cells
                    for c in range(min_col, max_col + 1):
                        val_by_col[c] += max(1, max_row - min_row + 1)

        errs = 0
        scanned = 0
        for row in ws.iter_rows(min_row=1, max_row=used_rows, min_col=1, max_col=used_cols):
            for cell in row:
                scanned += 1
                if cell.data_type == "e":
                    errs += 1

        fmt[ws.title] = {
            "cellsScanned": scanned,
            "validations": int(val_count),
            "errorsInCells": int(errs),
            "validationsByCol": {str(k): int(v) for k, v in val_by_col.items()},
        }

        header = [ws.cell(row=1, column=c).value for c in range(1, min(used_cols, 25) + 1)]
        sample = []
        for r in range(2, min(used_rows, 6) + 1):
            sample.append([ws.cell(row=r, column=c).value for c in range(1, min(used_cols, 25) + 1)])
        vprev[ws.title] = {"usedRange": dim, "usedRows": used_rows, "usedCols": used_cols, "header": header, "sampleRows": sample}

        if ws.title == "CONCURSOS":
            for r in ws.iter_rows(min_row=1, max_row=used_rows, min_col=1, max_col=used_cols, values_only=True):
                concursos_rows.append(list(r))

    # mismos checks de calidad de datos sobre CONCURSOS
    if not concursos_rows:
        findings.append({"severity": "CRITICAL", "area": "SHEET_DATA", "title": "CONCURSOS vacío/no legible",
                         "detail": "No hay matriz de datos disponible", "evidence": {}, "recommendation": "Revisar pestaña/permisos."})
    else:
        h = [norm(str(x) if x is not None else "") for x in concursos_rows[0]]
        rows = concursos_rows[1:]
        hm = {h[i]: i for i in range(len(h))}
        exp = ["NOMBRE CONCURSO", "ESTADO", "INSCRIPCION", "FECHA LIMITE INSCRIPCION", "EMAIL"]
        miss = [x for x in exp if x not in hm]
        if miss:
            findings.append({"severity": "HIGH", "area": "SHEET_STRUCTURE", "title": "Cabeceras ausentes",
                             "detail": "Faltan columnas clave", "evidence": {"missing": miss}, "recommendation": "Alinear cabeceras."})
        i_nom = hm.get("NOMBRE CONCURSO", 0); i_est = hm.get("ESTADO", 1); i_ins = hm.get("INSCRIPCION", 2)
        i_fli = hm.get("FECHA LIMITE INSCRIPCION", 3); i_ema = hm.get("EMAIL", 14)
        i_l1 = hm.get("LINK BASES/FORMULARIO", 11); i_l2 = hm.get("WEB / BASES / FORMULARIO ANO ANTERIOR", 12); i_l3 = hm.get("WEB / BASES / FORMULARIO 2 ANOS ANTERIOR", 13)
        ok_est = {"REVISAR", "REVISADO IA", "REVISADO HUMANO", "NUEVO DESCUBRIMIENTO"}
        ok_ins = {"ABIERTA", "CERRADA", "SIN PUBLICAR"}
        bad_est = []; bad_ins = []; bad_mail = []; bad_link = []; bad_name = []; bad_date = []
        dups = collections.defaultdict(list)
        for rn, row in enumerate(rows, start=2):
            v = lambda i: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
            nom = v(i_nom); est = norm(v(i_est)); ins = norm(v(i_ins)); fl = v(i_fli); em = v(i_ema).lower()
            if not nom and any((str(x).strip() if x is not None else "") for x in row):
                bad_name.append(rn)
            if nom:
                dups[norm(nom)].append(rn)
            if est and est not in ok_est:
                bad_est.append({"row": rn, "value": est})
            if ins and ins not in ok_ins:
                bad_ins.append({"row": rn, "value": ins})
            if em and "no publicado" not in em and "no localizado" not in em and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", em):
                bad_mail.append({"row": rn, "value": em})
            for lk in [v(i_l1), v(i_l2), v(i_l3)]:
                if lk and "NO PUBLICADO" not in norm(lk) and "NO LOCALIZADO" not in norm(lk) and not re.match(r"^https?://", lk, re.I):
                    bad_link.append({"row": rn, "value": lk})
            if fl and "ESTIMADO" not in norm(fl):
                if not (re.match(r"^\d+(\.\d+)?$", fl) or re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", fl) or re.search(r"\d{4}-\d{1,2}-\d{1,2}", fl)):
                    bad_date.append({"row": rn, "value": fl})
        dup_groups = [{"nameNorm": k, "rows": v} for k, v in dups.items() if k and len(v) > 1]
        if bad_est:
            findings.append({"severity": "HIGH", "area": "SHEET_DATA", "title": "ESTADO inválido",
                             "detail": f"{len(bad_est)} filas", "evidence": {"examples": bad_est[:15]}, "recommendation": "Corregir catálogo ESTADO."})
        if bad_ins:
            findings.append({"severity": "HIGH", "area": "SHEET_DATA", "title": "INSCRIPCIÓN inválida",
                             "detail": f"{len(bad_ins)} filas", "evidence": {"examples": bad_ins[:15]}, "recommendation": "Corregir catálogo INSCRIPCIÓN."})
        if bad_mail:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Email sospechoso",
                             "detail": f"{len(bad_mail)} filas", "evidence": {"examples": bad_mail[:15]}, "recommendation": "Revisar formato email."})
        if bad_link:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Links inválidos",
                             "detail": f"{len(bad_link)} filas", "evidence": {"examples": bad_link[:15]}, "recommendation": "Añadir https://."})
        if bad_name:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Filas sin nombre",
                             "detail": f"{len(bad_name)} filas", "evidence": {"rows": bad_name[:30]}, "recommendation": "Completar o limpiar filas."})
        if dup_groups:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Posibles duplicados",
                             "detail": f"{len(dup_groups)} grupos", "evidence": {"examples": dup_groups[:15]}, "recommendation": "Depurar duplicados."})
        if bad_date:
            findings.append({"severity": "MEDIUM", "area": "SHEET_DATA", "title": "Fecha límite no parseable",
                             "detail": f"{len(bad_date)} filas", "evidence": {"examples": bad_date[:15]}, "recommendation": "Estandarizar fecha."})
        dat_rows = max(vprev.get("CONCURSOS", {}).get("usedRows", 0) - 1, 0)
        vbc = {int(k): int(v) for k, v in fmt.get("CONCURSOS", {}).get("validationsByCol", {}).items()}
        weak = []
        for cname, idx in [("ESTADO", i_est + 1), ("INSCRIPCION", i_ins + 1)]:
            if dat_rows > 0 and vbc.get(idx, 0) < int(dat_rows * 0.8):
                weak.append({"column": cname, "index": idx, "validatedCells": vbc.get(idx, 0), "dataRows": dat_rows})
        if weak:
            findings.append({"severity": "HIGH", "area": "SHEET_VALIDATIONS", "title": "Validaciones incompletas",
                             "detail": "Columnas clave sin validación consistente", "evidence": {"columns": weak}, "recommendation": "Aplicar dropdown a toda la matriz."})

    err_sheets = [{"sheet": n, "errorsInCells": x["errorsInCells"]} for n, x in fmt.items() if int(x.get("errorsInCells", 0)) > 0]
    if err_sheets:
        findings.append({"severity": "HIGH", "area": "SHEET_FORMULAS", "title": "Errores de fórmula/valor",
                         "detail": "Hay celdas con error efectivo", "evidence": {"sheets": err_sheets}, "recommendation": "Corregir celdas con #ERROR/#REF."})
    total_prot = sum(feats[n]["protectedRanges"] for n in feats)
    if total_prot == 0:
        findings.append({"severity": "LOW", "area": "SHEET_PROTECTION", "title": "Sin protecciones",
                         "detail": "No hay rangos protegidos", "evidence": {"totalProtectedRanges": 0}, "recommendation": "Proteger celdas críticas."})

    overview = {
        "spreadsheet": {"id": spreadsheet_id, "title": title, "locale": "", "timeZone": "", "sheetCount": len(wb.worksheets)},
        "structure": structure,
        "sheetFeatures": feats,
        "formatsSummary": fmt,
        "valuesPreview": vprev,
        "fallback": {"mode": "xlsx_export", "path": tmp},
    }
    return overview, findings


def render_md(rep: dict) -> str:
    c = collections.Counter([f["severity"] for f in rep["findings"]])
    lines = [
        "# AUDITORIA EXTREMA CRM AYUDAS Y SUBVENCIONES",
        "",
        f"- Fecha: `{rep['meta']['timestamp']}`",
        f"- Spreadsheet: `{rep['context']['spreadsheetId']}`",
        f"- ScriptId: `{rep['context']['scriptId']}`",
        f"- READY: `{rep['context']['ready']}`",
        f"- Modo auditoría hoja: `{rep['context'].get('sheetAuditMode','')}`",
        "",
        "## Resumen",
        "",
        f"- CRITICAL: **{c.get('CRITICAL', 0)}**",
        f"- HIGH: **{c.get('HIGH', 0)}**",
        f"- MEDIUM: **{c.get('MEDIUM', 0)}**",
        f"- LOW: **{c.get('LOW', 0)}**",
        "",
        "## Hallazgos",
        "",
    ]
    if not rep["findings"]:
        lines.append("- Sin hallazgos.")
    else:
        order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        for f in sorted(rep["findings"], key=lambda x: (order.get(x["severity"], 9), x["area"], x["title"])):
            lines += [
                f"### [{f['severity']}] {f['title']}",
                f"- Area: `{f['area']}`",
                f"- Detalle: {f['detail']}",
                f"- Recomendacion: {f['recommendation']}",
                f"- Evidencia: `{json.dumps(f['evidence'], ensure_ascii=False)[:1200]}`",
                "",
            ]
    lines += ["## Cobertura", "", "- Estructura: SI", "- Datos/Formulas: SI", "- Formatos: SI", "- Combinadas: SI", "- Validaciones: SI", "- Filtros/vistas: SI", "- Formato condicional: SI", "- Protecciones: SI", "", "## Pestañas", ""]
    for s in rep["sheetOverview"]["structure"]:
        lines.append(f"- `{s['title']}` | used `{s['usedRange']}` | usados `{s['usedRows']}x{s['usedCols']}` | config `{s['rowsConfigured']}x{s['colsConfigured']}`")
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--spreadsheet-id", default="1LgZG2ObSCJzEQvrysu1NFFEvYlupLXVByDnIMCr-wYA")
    ap.add_argument("--project-dir", default=PROJ)
    ap.add_argument("--clasprc", default=CLASPRC)
    ap.add_argument("--reports-dir", default=REPORTS)
    ap.add_argument("--script-id", default="")
    args = ap.parse_args()

    os.makedirs(args.reports_dir, exist_ok=True)
    clasp = load_json(os.path.join(args.project_dir, ".clasp.json"))
    script_id = args.script_id or clasp.get("scriptId", "")
    if not script_id:
        raise RuntimeError("scriptId no disponible")
    token = refresh_token(args.clasprc)
    sm = script_project(token, script_id)
    ready = sm.get("parentId") == args.spreadsheet_id
    sheet_mode = "sheets_api"
    try:
        sh_over, sh_find = audit_sheet(token, args.spreadsheet_id)
    except Exception as e:
        sheet_mode = "xlsx_fallback"
        sh_over, sh_find = audit_sheet_xlsx_fallback(token, args.spreadsheet_id)
        sh_find.append({
            "severity": "LOW",
            "area": "SHEET_AUDIT_MODE",
            "title": "Auditoría en modo fallback XLSX",
            "detail": "Google Sheets API no disponible; se usó exportación XLSX por Drive API.",
            "evidence": {"error": str(e)[:1200]},
            "recommendation": "Si quieres inspección con metadato nativo completo, habilitar Sheets API en el proyecto OAuth.",
        })
    code_over, code_find = audit_code(args.project_dir, token, script_id)
    report = {
        "meta": {"timestamp": dt.datetime.now().isoformat(timespec="seconds"), "tool": "auditar_crm_ayudas_extremo.py"},
        "context": {"spreadsheetId": args.spreadsheet_id, "scriptId": script_id, "scriptTitle": sm.get("title", ""), "scriptParentId": sm.get("parentId", ""), "ready": ready, "projectDir": args.project_dir, "sheetAuditMode": sheet_mode},
        "sheetOverview": sh_over,
        "codeOverview": code_over,
        "findings": sh_find + code_find,
    }
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    jpath = os.path.join(args.reports_dir, f"audit_crm_ayudas_{stamp}.json")
    mpath = os.path.join(args.reports_dir, f"audit_crm_ayudas_{stamp}.md")
    with open(jpath, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    with open(mpath, "w", encoding="utf-8") as f:
        f.write(render_md(report))
    print(json.dumps({"json": jpath, "md": mpath, "findings": len(report["findings"])}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
