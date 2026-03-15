#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import tempfile
import urllib.parse
import urllib.request
import urllib.error

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
CLASPRC = os.path.join(os.environ.get("USERPROFILE", ""), ".clasprc.json")


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def http_json(method: str, url: str, token: str | None = None, data: dict | None = None) -> dict:
    headers = {"Accept": "application/json"}
    payload = None
    if token:
        headers["Authorization"] = "Bearer " + token
    if data is not None:
        payload = json.dumps(data).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url=url, method=method, headers=headers, data=payload)
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url} :: {body}") from e


def http_bytes(method: str, url: str, token: str | None = None) -> bytes:
    headers = {"Accept": "*/*"}
    if token:
        headers["Authorization"] = "Bearer " + token
    req = urllib.request.Request(url=url, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            return r.read()
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url} :: {body}") from e


def get_token_from_profile(clasprc: str, profile: str) -> str:
    cfg = load_json(clasprc)
    if profile and cfg.get("tokens", {}).get(profile):
        token = str(cfg["tokens"][profile].get("access_token", ""))
        if token:
            return token
    if cfg.get("tokens", {}).get("default"):
        token = str(cfg["tokens"]["default"].get("access_token", ""))
        if token:
            return token
    return str(cfg.get("token", {}).get("access_token", ""))


def refresh_token(clasprc: str, profile: str) -> str:
    cfg = load_json(clasprc)
    client_id = ""
    client_secret = ""
    refresh = ""
    if cfg.get("token") and cfg.get("oauth2ClientSettings"):
        client_id = str(cfg["oauth2ClientSettings"].get("clientId", ""))
        client_secret = str(cfg["oauth2ClientSettings"].get("clientSecret", ""))
        refresh = str(cfg["token"].get("refresh_token", ""))
    elif cfg.get("tokens", {}).get("default"):
        token = cfg["tokens"]["default"]
        client_id = str(token.get("client_id", ""))
        client_secret = str(token.get("client_secret", ""))
        refresh = str(token.get("refresh_token", ""))
    if not (client_id and client_secret and refresh):
        raise RuntimeError("No se pudo leer OAuth desde .clasprc.json")
    try:
        out = http_json(
            "POST",
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh,
                "grant_type": "refresh_token",
            },
        )
        access = out.get("access_token", "")
        if access:
            return access
    except Exception:
        pass
    fallback = get_token_from_profile(clasprc, profile)
    if fallback:
        return fallback
    raise RuntimeError("No se pudo renovar OAuth ni recuperar token activo del perfil")


def export_sheet_xlsx(token: str, spreadsheet_id: str, out_path: str) -> str:
    url = (
        f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}/export"
        "?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    data = http_bytes("GET", url, token)
    with open(out_path, "wb") as f:
        f.write(data)
    return out_path


def a1_range(max_row: int, max_col: int) -> str:
    if max_row <= 0 or max_col <= 0:
        return "A1"
    return "A1:" + get_column_letter(max_col) + str(max_row)


def clean(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def make_jsonable(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, list):
        return [make_jsonable(v) for v in value]
    if isinstance(value, dict):
        return {str(k): make_jsonable(v) for k, v in value.items()}
    return value


def preview_rows(ws, max_cols: int, start: int, end: int):
    rows = []
    for r in range(start, min(end, ws.max_row) + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)])
    return rows


def collect_format_summary(ws, used_rows: int, used_cols: int) -> dict:
    unique = {
        "fills": set(),
        "fonts": set(),
        "sizes": set(),
        "bold": set(),
        "italic": set(),
        "align_h": set(),
        "align_v": set(),
        "numfmt": set(),
        "borders": set(),
    }

    scan_rows = min(used_rows, 400)
    scan_cols = min(used_cols, 30)
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            font = cell.font
            align = cell.alignment
            border = cell.border
            unique["fills"].add(str(getattr(fill.fgColor, "rgb", "") or getattr(fill.fgColor, "index", "")))
            unique["fonts"].add(clean(font.name))
            unique["sizes"].add(str(font.sz))
            unique["bold"].add(str(bool(font.bold)))
            unique["italic"].add(str(bool(font.italic)))
            unique["align_h"].add(clean(align.horizontal))
            unique["align_v"].add(clean(align.vertical))
            unique["numfmt"].add(clean(cell.number_format))
            unique["borders"].add(
                "|".join(
                    [
                        clean(getattr(border.left, "style", "")),
                        clean(getattr(border.right, "style", "")),
                        clean(getattr(border.top, "style", "")),
                        clean(getattr(border.bottom, "style", "")),
                    ]
                )
            )

    return {
        "fills": sorted([x for x in unique["fills"] if x])[:60],
        "fonts": sorted([x for x in unique["fonts"] if x])[:30],
        "sizes": sorted([x for x in unique["sizes"] if x])[:30],
        "bold": sorted([x for x in unique["bold"] if x])[:5],
        "italic": sorted([x for x in unique["italic"] if x])[:5],
        "align_h": sorted([x for x in unique["align_h"] if x])[:20],
        "align_v": sorted([x for x in unique["align_v"] if x])[:20],
        "numfmt": sorted([x for x in unique["numfmt"] if x])[:80],
        "borders": sorted([x for x in unique["borders"] if x])[:25],
    }


def inspect_sheet(ws) -> dict:
    used_rows = int(ws.max_row or 0)
    used_cols = int(ws.max_column or 0)
    dim = ws.calculate_dimension()
    non_empty_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=used_rows, min_col=1, max_col=used_cols, values_only=True):
        if any(clean(v) for v in row):
            non_empty_rows += 1

    formulas = []
    errors = 0
    scan_rows = min(used_rows, 1500)
    scan_cols = min(used_cols, 40)
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.data_type == "f":
                formulas.append({"cell": cell.coordinate, "formula": clean(cell.value)})
            if cell.data_type == "e":
                errors += 1

    validations = []
    validations_by_col = {}
    total_validations = 0
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            refs = str(dv.sqref).split()
            validations.append(
                {
                    "type": clean(dv.type),
                    "allowBlank": bool(getattr(dv, "allowBlank", False)),
                    "operator": clean(getattr(dv, "operator", "")),
                    "formula1": clean(getattr(dv, "formula1", "")),
                    "formula2": clean(getattr(dv, "formula2", "")),
                    "ranges": refs[:8],
                }
            )
            for ref in refs:
                m = re.match(r"([A-Z]+)", ref)
                if m:
                    col = m.group(1)
                    validations_by_col[col] = validations_by_col.get(col, 0) + 1
                total_validations += 1

    cond = []
    try:
        for key, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                cond.append({"range": clean(str(key)), "type": clean(getattr(rule, "type", ""))})
    except Exception:
        cond = []

    format_summary = collect_format_summary(ws, used_rows, used_cols)
    max_cols_preview = min(max(used_cols, 1), 20)
    return {
        "name": ws.title,
        "hidden": ws.sheet_state != "visible",
        "dimensions": {
            "usedRows": used_rows,
            "usedCols": used_cols,
            "usedRange": dim if dim else a1_range(used_rows, used_cols),
            "nonEmptyRows": non_empty_rows,
        },
        "visibleData": {
            "header": preview_rows(ws, max_cols_preview, 1, 1)[0] if used_rows else [],
            "sampleRows": preview_rows(ws, max_cols_preview, 2, 6),
        },
        "formulas": {"total": len(formulas), "sample": formulas[:120], "errorsInCells": errors},
        "formatting": format_summary,
        "merges": [str(rg) for rg in ws.merged_cells.ranges],
        "validations": {
            "total": total_validations,
            "byColumn": validations_by_col,
            "sample": validations[:80],
        },
        "filters": {
            "basicFilterRange": clean(getattr(ws.auto_filter, "ref", "")),
            "filterViews": [],
        },
        "conditionalFormatting": {"total": len(cond), "rules": cond[:120]},
        "protections": {
            "sheetProtected": bool(ws.protection and ws.protection.sheet),
            "rangeProtected": "No disponible por exportacion XLSX",
        },
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--spreadsheet-id", default="")
    ap.add_argument("--xlsx-path", default="")
    ap.add_argument("--clasprc", default=CLASPRC)
    ap.add_argument("--token-profile", default="default")
    ap.add_argument("--reports-dir", default=REPORTS)
    args = ap.parse_args()

    os.makedirs(args.reports_dir, exist_ok=True)
    spreadsheet_id = clean(args.spreadsheet_id)
    xlsx_path = clean(args.xlsx_path)
    meta = {
        "id": spreadsheet_id or "LOCAL_XLSX",
        "name": os.path.basename(xlsx_path) if xlsx_path else "",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "modifiedTime": "",
        "owners": [],
    }

    if xlsx_path:
        tmp_path = xlsx_path
    else:
        if not spreadsheet_id:
            raise RuntimeError("Debes indicar --spreadsheet-id o --xlsx-path")
        token = refresh_token(args.clasprc, args.token_profile)
        meta = http_json(
            "GET",
            "https://www.googleapis.com/drive/v3/files/"
            + spreadsheet_id
            + "?fields=id,name,mimeType,modifiedTime,owners(displayName,emailAddress)",
            token,
        )
        tmp_path = os.path.join(tempfile.gettempdir(), f"crm_sheet_fallback_{spreadsheet_id}.xlsx")
        export_sheet_xlsx(token, spreadsheet_id, tmp_path)

    wb = load_workbook(tmp_path, data_only=False)

    report = {
        "meta": {
            "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
            "mode": "xlsx_fallback",
            "note": "Inspeccion por exportacion XLSX porque no hay API executable disponible",
        },
        "workbook": {
            "id": meta.get("id", args.spreadsheet_id),
            "name": meta.get("name", ""),
            "mimeType": meta.get("mimeType", ""),
            "modifiedTime": meta.get("modifiedTime", ""),
            "owners": meta.get("owners", []),
            "sheetCount": len(wb.worksheets),
            "xlsxPath": tmp_path,
        },
        "sheets": [inspect_sheet(ws) for ws in wb.worksheets],
    }

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    sid_for_name = spreadsheet_id if spreadsheet_id else "local_xlsx"
    json_path = os.path.join(args.reports_dir, f"audit_sheet_fallback_{sid_for_name}_{stamp}.json")
    md_path = os.path.join(args.reports_dir, f"audit_sheet_fallback_{sid_for_name}_{stamp}.md")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(make_jsonable(report), f, ensure_ascii=False, indent=2)

    lines = [
        "# AUDITORIA SHEET (FALLBACK XLSX)",
        "",
        f"- Fecha: {report['meta']['generatedAt']}",
        f"- Spreadsheet: {report['workbook']['id']}",
        f"- Nombre: {report['workbook']['name']}",
        f"- Hojas: {report['workbook']['sheetCount']}",
        "",
        "## Cobertura",
        "- Estructura: SI",
        "- Datos visibles y formulas: SI",
        "- Formatos: SI",
        "- Celdas combinadas: SI",
        "- Validaciones/desplegables: SI",
        "- Filtros/vistas filtro: SI (basic filter; filter views no exportable)",
        "- Formato condicional: SI",
        "- Protecciones: SI (sheet; rango detallado no exportable)",
        "",
        "## Hojas",
    ]
    for sh in report["sheets"]:
        dim = sh["dimensions"]
        lines.append(
            f"- {sh['name']} | rango {dim['usedRange']} | usadas {dim['usedRows']}x{dim['usedCols']} | no vacias {dim['nonEmptyRows']}"
        )
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(json.dumps({"json": json_path, "md": md_path, "sheetCount": len(report["sheets"])}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
